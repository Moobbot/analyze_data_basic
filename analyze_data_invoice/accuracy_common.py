#!/usr/bin/env python3
"""
Common utilities for invoice accuracy evaluation
"""

import pandas as pd
import re
from openpyxl.styles import Font, PatternFill


# Fields to evaluate
EVAL_FIELDS = [
    "Type",
    "No",
    "Date",
    "Customer",
    "Supplier",
    "Currency",
    "Ex rate",
    "Ex rate to SGD",
    "Tax type",
    "Description",
    "Amount (before tax)",
    "Tax amount",
    "Amount (after GST)",
    "Amount in SGD",
    "Tax amount in SGD",
    "Amount after tax in SGD",
]


def normalize_invoice_name(name):
    """Normalize invoice name for matching"""
    if pd.isna(name):
        return ""
    name = str(name).strip()

    # Remove .pdf extension if present
    if name.lower().endswith(".pdf"):
        name = name[:-4]

    # Remove [0], [1], [2] etc. suffixes
    name = re.sub(r"\[\d+\]$", "", name)

    return name.strip()


def normalize_value(value):
    """Normalize value for comparison"""
    if pd.isna(value):
        return ""

    value_str = str(value).strip()

    # Try to parse as number for numeric comparison
    try:
        # Remove commas and parse as float
        num_val = float(value_str.replace(",", ""))
        # Return normalized number string (removes trailing zeros)
        return str(num_val).lower()
    except (ValueError, TypeError):
        # Not a number, treat as string
        value_str = value_str.replace(",", "")
        return value_str.lower()


def normalize_company_name(name):
    """Normalize company name for fuzzy matching"""
    if pd.isna(name) or str(name).strip() == "":
        return ""

    name = str(name).strip().lower()

    # Remove parenthetical suffixes (e.g., "(SGD)", "(USD)", "(Sponsor of...)")
    # This allows "Abadi Investments Pte Ltd" to match "Abadi Investments Pte Ltd (SGD)"
    name = re.sub(r"\s*\([^)]*\)\s*$", "", name).strip()

    # Common abbreviations mapping
    replacements = {
        r"\bltd\.?\b": "limited",
        r"\bpte\.?\b": "private",
        r"\bco\.?\b": "company",
        r"\bcorp\.?\b": "corporation",
        r"\binc\.?\b": "incorporated",
        r"\bllc\.?\b": "limited liability company",
        r"\bllp\.?\b": "limited liability partnership",
        r"\bs\.?a\.?\b": "sociedad anonima",
        r"\bspc\.?\b": "segregated portfolio company",
    }

    for pattern, replacement in replacements.items():
        name = re.sub(pattern, replacement, name)

    # Remove extra spaces
    name = re.sub(r"\s+", " ", name).strip()

    return name


def normalize_tax_type(tax_str):
    """Normalize tax type to extract percentage"""
    if pd.isna(tax_str) or str(tax_str).strip() == "":
        return ""

    tax_str = str(tax_str).strip().lower()

    # Extract percentage
    match = re.search(r"(\d+(?:\.\d+)?)\s*%", tax_str)
    if match:
        return f"{match.group(1)}%"

    # Common tax keywords
    if "exempt" in tax_str or "zero" in tax_str or "nil" in tax_str:
        return "0%"

    # Return normalized string if no percentage found
    return tax_str


def are_values_equivalent(val1, val2, field_name=None):
    """Check if two values are equivalent (handles numeric and string comparison)
    Returns: (is_match, match_type)
    - match_type: 'exact', 'numeric', 'mismatch'
    """
    # Handle empty values
    if pd.isna(val1) and pd.isna(val2):
        return True, "exact"

    str1 = str(val1).strip() if not pd.isna(val1) else ""
    str2 = str(val2).strip() if not pd.isna(val2) else ""

    if str1 == "" and str2 == "":
        return True, "exact"

    # Special case: Treat "" as "0" or "0.0"
    # Check if value is zero in any format
    def is_zero_value(s):
        if s == "":
            return True
        # Try to convert to float and check if it's zero
        try:
            num = float(s.replace(",", ""))  # Remove thousand separators
            return num == 0.0 or num == -0.0
        except (ValueError, TypeError):
            # If can't convert to number, check string patterns
            cleaned = (
                s.replace(".", "").replace(",", "").replace(" ", "").replace("-", "")
            )
            return cleaned == "" or cleaned == "0"

    is_zero1 = is_zero_value(str1)
    is_zero2 = is_zero_value(str2)

    if is_zero1 and is_zero2:
        # Check if one is actually empty and the other is explicitly 0
        if (str1 == "" and str2 != "") or (str1 != "" and str2 == ""):
            return True, "zero_equivalent"
        # Both are 0 but different format (e.g., "0" vs "0.0")
        if str1 != str2:
            return True, "zero_equivalent"
        return True, "exact"

    if str1 == "" or str2 == "":
        return False, "mismatch"

    # Apply field-specific fuzzy matching
    if field_name in ["Customer", "Supplier"]:
        norm1 = normalize_company_name(val1)
        norm2 = normalize_company_name(val2)
        if norm1 == norm2:
            # Check if original strings were different
            if str1.lower() != str2.lower():
                return True, "fuzzy"
            return True, "exact"

    if field_name == "Tax type":
        norm1 = normalize_tax_type(val1)
        norm2 = normalize_tax_type(val2)
        if norm1 == norm2:
            # Check if original strings were different
            if str1.lower() != str2.lower():
                return True, "fuzzy"
            return True, "exact"

    # Normalize both values
    norm1 = normalize_value(val1)
    norm2 = normalize_value(val2)

    # Direct string match after normalization
    if norm1 == norm2:
        # Check if original strings were different (different format)
        orig1 = str(val1).strip()
        orig2 = str(val2).strip()
        if orig1 != orig2:
            return True, "numeric"
        return True, "exact"

    return False, "mismatch"


def safe_float(value):
    """Safely convert value to float"""
    try:
        if pd.isna(value):
            return None
        return float(str(value).replace(",", ""))
    except (ValueError, TypeError):
        return None


def get_tax_rate(tax_type):
    """Get tax rate from tax type string"""
    if pd.isna(tax_type):
        return 0.0

    tax_str = str(tax_type).lower()

    # Extract percentage from strings like "7%", "GST 7%", etc.
    match = re.search(r"(\d+(?:\.\d+)?)\s*%", tax_str)
    if match:
        return float(match.group(1)) / 100

    # Common tax rates
    if "7" in tax_str:
        return 0.07
    if "9" in tax_str:
        return 0.09
    if "exempt" in tax_str or "zero" in tax_str:
        return 0.0

    return 0.0


def calculate_missing_fields(row):
    """Calculate missing SGD and tax amount fields if not present"""
    calculated_fields = {}

    # Get exchange rate
    exrate = safe_float(row.get("Ex rate to SGD"))
    if exrate is None:
        exrate = safe_float(row.get("Ex rate"))

    # Calculate Amount in SGD if missing
    if pd.isna(row.get("Amount in SGD")) or str(row.get("Amount in SGD")).strip() == "":
        amount_before_tax = safe_float(row.get("Amount (before tax)"))
        if amount_before_tax is not None and exrate is not None:
            calculated_fields["Amount in SGD"] = round(amount_before_tax * exrate, 12)

    # Calculate Tax amount in SGD if missing
    if (
        pd.isna(row.get("Tax amount in SGD"))
        or str(row.get("Tax amount in SGD")).strip() == ""
    ):
        tax_amount = safe_float(row.get("Tax amount"))
        if tax_amount is not None and exrate is not None:
            calculated_fields["Tax amount in SGD"] = round(tax_amount * exrate, 12)

    # Calculate Amount after tax in SGD if missing
    if (
        pd.isna(row.get("Amount after tax in SGD"))
        or str(row.get("Amount after tax in SGD")).strip() == ""
    ):
        amount_after_gst = safe_float(row.get("Amount (after GST)"))
        if amount_after_gst is not None and exrate is not None:
            calculated_fields["Amount after tax in SGD"] = round(
                amount_after_gst * exrate, 12
            )

    # Calculate Tax amount if missing
    if pd.isna(row.get("Tax amount")) or str(row.get("Tax amount")).strip() == "":
        amount_before_tax = safe_float(row.get("Amount (before tax)"))
        tax_rate = get_tax_rate(row.get("Tax type"))
        if amount_before_tax is not None and tax_rate > 0:
            calculated_fields["Tax amount"] = round(amount_before_tax * tax_rate, 12)

    return calculated_fields


def calculate_field_accuracy(gt_df, model_df, field):
    """Calculate accuracy metrics for a specific field"""

    # Get matched invoices only
    gt_invoices = set(gt_df["invoice_name_normalized"].unique())
    model_invoices = set(model_df["invoice_name_normalized"].unique())
    matched_invoices = gt_invoices & model_invoices

    # Filter to matched invoices
    gt_matched = gt_df[gt_df["invoice_name_normalized"].isin(matched_invoices)]
    model_matched = model_df[model_df["invoice_name_normalized"].isin(matched_invoices)]

    # Count matches
    total_gt = 0
    total_model = 0
    correct = 0

    for invoice_name in matched_invoices:
        # Only consider rows for this specific invoice
        gt_rows = gt_matched[gt_matched["invoice_name_normalized"] == invoice_name]
        model_rows = model_matched[
            model_matched["invoice_name_normalized"] == invoice_name
        ]

        total_gt += len(gt_rows)
        total_model += len(model_rows)

        for i in range(min(len(gt_rows), len(model_rows))):
            gt_val = gt_rows.iloc[i].get(field)
            model_val = model_rows.iloc[i].get(field)

            is_match, _ = are_values_equivalent(gt_val, model_val, field)
            if is_match:
                correct += 1

    # Calculate metrics
    accuracy = (correct / total_gt * 100) if total_gt > 0 else 0
    precision = (correct / total_model * 100) if total_model > 0 else 0
    recall = (correct / total_gt * 100) if total_gt > 0 else 0
    f1_score = (
        (2 * precision * recall / (precision + recall))
        if (precision + recall) > 0
        else 0
    )

    return {
        "field": field,
        "total_gt": total_gt,
        "total_model": total_model,
        "correct": correct,
        "accuracy": accuracy,
        "precision": precision,
        "recall": recall,
        "f1_score": f1_score,
    }


def create_excel_report(wb, gt_df, model_df, results, matched_invoices):
    """Create Excel sheets with accuracy summary, field comparison AND Type Analysis"""

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Styling
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True)
    match_fill = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )
    mismatch_fill = PatternFill(
        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
    )
    calculated_fill = PatternFill(
        start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
    )
    numeric_match_fill = PatternFill(
        start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
    )
    zero_match_fill = PatternFill(
        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
    )  # Light blue for zero-equivalent
    fuzzy_match_fill = PatternFill(
        start_color="FFE6CC", end_color="FFE6CC", fill_type="solid"
    )  # Light orange for fuzzy match

    # Sheet 1: Accuracy Summary
    ws_summary = wb.create_sheet("Accuracy Summary")

    headers = [
        "Field",
        "Total GT",
        "Total Model",
        "Correct",
        "Accuracy (%)",
        "Precision (%)",
        "Recall (%)",
        "F1-Score (%)",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws_summary.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, result in enumerate(results, start=2):
        ws_summary.cell(row_idx, 1, result["field"])
        ws_summary.cell(row_idx, 2, result["total_gt"])
        ws_summary.cell(row_idx, 3, result["total_model"])
        ws_summary.cell(row_idx, 4, result["correct"])
        ws_summary.cell(row_idx, 5, f"{result['accuracy']:.2f}")
        ws_summary.cell(row_idx, 6, f"{result['precision']:.2f}")
        ws_summary.cell(row_idx, 7, f"{result['recall']:.2f}")
        ws_summary.cell(row_idx, 8, f"{result['f1_score']:.2f}")

    # Overall Summary
    summary_row = len(results) + 3
    ws_summary.cell(summary_row, 1, "OVERALL SUMMARY").font = Font(bold=True)

    total_gt = sum(r["total_gt"] for r in results)
    total_model = sum(r["total_model"] for r in results)
    total_correct = sum(r["correct"] for r in results)

    overall_accuracy = (total_correct / total_gt * 100) if total_gt > 0 else 0
    overall_precision = (total_correct / total_model * 100) if total_model > 0 else 0
    overall_recall = (total_correct / total_gt * 100) if total_gt > 0 else 0
    overall_f1 = (
        (2 * overall_precision * overall_recall / (overall_precision + overall_recall))
        if (overall_precision + overall_recall) > 0
        else 0
    )

    ws_summary.cell(summary_row, 2, total_gt)
    ws_summary.cell(summary_row, 3, total_model)
    ws_summary.cell(summary_row, 4, total_correct)
    ws_summary.cell(summary_row, 5, f"{overall_accuracy:.2f}")
    ws_summary.cell(summary_row, 6, f"{overall_precision:.2f}")
    ws_summary.cell(summary_row, 7, f"{overall_recall:.2f}")
    ws_summary.cell(summary_row, 8, f"{overall_f1:.2f}")

    ws_summary.column_dimensions["A"].width = 25
    for col in ["B", "C", "D", "E", "F", "G", "H"]:
        ws_summary.column_dimensions[col].width = 15

    # Sheet 2: Type Analysis (New)
    if "invoice_type" in gt_df.columns:
        ws_type = wb.create_sheet("Type Analysis")
        type_headers = [
            "Invoice Type",
            "Matched Invoices",
            "Total GT Rows",
            "Total Model Rows",
            "Accuracy (%)",
            "Precision (%)",
            "Recall (%)",
            "F1-Score (%)",
        ]

        for col, header in enumerate(type_headers, start=1):
            cell = ws_type.cell(1, col, header)
            cell.fill = header_fill
            cell.font = header_font

        # Calculate metrics per type
        invoice_types = gt_df["invoice_type"].unique()
        type_results = []

        for inv_type in sorted([t for t in invoice_types if pd.notna(t)]):
            # Filter DataFrames for this type
            type_gt_df = gt_df[gt_df["invoice_type"] == inv_type]

            # Find matching invoices in model output that belong to this type
            # Note: Model output might not have 'invoice_type' populated reliably if we didn't map it.
            # Strategy: Use invoices from type_gt_df and find them in model_df
            type_invoices = set(type_gt_df["invoice_name_normalized"].unique())
            type_model_df = model_df[
                model_df["invoice_name_normalized"].isin(type_invoices)
            ]

            # Calculate overall metrics for this type (aggregating all fields)
            type_correct = 0
            type_total_gt = 0
            type_total_model = 0

            # Matched invoices for this type
            matched_type_invoices = type_invoices.intersection(
                set(type_model_df["invoice_name_normalized"].unique())
            )

            for field in EVAL_FIELDS:
                metrics = calculate_field_accuracy(type_gt_df, type_model_df, field)
                type_correct += metrics["correct"]
                type_total_gt += metrics["total_gt"]
                type_total_model += metrics["total_model"]

            # Calculate aggregate metrics
            acc = (type_correct / type_total_gt * 100) if type_total_gt > 0 else 0
            prec = (
                (type_correct / type_total_model * 100) if type_total_model > 0 else 0
            )
            rec = (type_correct / type_total_gt * 100) if type_total_gt > 0 else 0
            f1 = (2 * prec * rec / (prec + rec)) if (prec + rec) > 0 else 0

            type_results.append(
                {
                    "type": inv_type,
                    "matched_invoices": len(matched_type_invoices),
                    "total_gt": type_total_gt,
                    "total_model": type_total_model,
                    "accuracy": acc,
                    "precision": prec,
                    "recall": rec,
                    "f1": f1,
                }
            )

        # Write to sheet
        for row_idx, res in enumerate(type_results, start=2):
            ws_type.cell(row_idx, 1, res["type"])
            ws_type.cell(row_idx, 2, res["matched_invoices"])
            ws_type.cell(row_idx, 3, res["total_gt"])
            ws_type.cell(row_idx, 4, res["total_model"])
            ws_type.cell(row_idx, 5, f"{res['accuracy']:.2f}")
            ws_type.cell(row_idx, 6, f"{res['precision']:.2f}")
            ws_type.cell(row_idx, 7, f"{res['recall']:.2f}")
            ws_type.cell(row_idx, 8, f"{res['f1']:.2f}")

        ws_type.column_dimensions["A"].width = 20
        for col in range(2, 9):
            ws_type.column_dimensions[chr(64 + col)].width = 15

    # Sheet 3: Detailed Comparison (Existing)
    ws_compare = wb.create_sheet("Field Comparison")

    col = 1
    ws_compare.cell(1, col, "Invoice Name").fill = header_fill
    ws_compare.cell(1, col, "Invoice Name").font = header_font
    col += 1
    if "invoice_type" in gt_df.columns:
        ws_compare.cell(1, col, "Type").fill = header_fill
        ws_compare.cell(1, col, "Type").font = header_font
        col += 1

    for field in EVAL_FIELDS:
        ws_compare.cell(1, col, f"GT_{field}").fill = header_fill
        ws_compare.cell(1, col, f"GT_{field}").font = header_font
        ws_compare.cell(1, col + 1, f"Model_{field}").fill = header_fill
        ws_compare.cell(1, col + 1, f"Model_{field}").font = header_font
        ws_compare.cell(1, col + 2, "Match").fill = header_fill
        ws_compare.cell(1, col + 2, "Match").font = header_font
        col += 3

    row = 2
    for invoice_name in sorted(matched_invoices):
        gt_rows = gt_df[gt_df["invoice_name_normalized"] == invoice_name]
        model_rows = model_df[model_df["invoice_name_normalized"] == invoice_name]

        # Get invoice type
        inv_type = gt_rows.iloc[0].get("invoice_type", "") if not gt_rows.empty else ""

        max_rows = max(len(gt_rows), len(model_rows))

        for i in range(max_rows):
            col = 1
            ws_compare.cell(row, col, invoice_name)
            col += 1
            if "invoice_type" in gt_df.columns:
                ws_compare.cell(row, col, inv_type)
                col += 1

            # Get data
            gt_row_data = gt_rows.iloc[i].to_dict() if i < len(gt_rows) else {}
            model_row_data = model_rows.iloc[i].to_dict() if i < len(model_rows) else {}

            calculated_gt = calculate_missing_fields(gt_row_data) if gt_row_data else {}

            for field in EVAL_FIELDS:
                gt_val = gt_row_data.get(field) if gt_row_data else ""
                is_calculated = False

                if field in calculated_gt and (
                    pd.isna(gt_val) or str(gt_val).strip() == ""
                ):
                    gt_val = calculated_gt[field]
                    is_calculated = True

                model_val = model_row_data.get(field) if model_row_data else ""

                gt_str = (
                    str(gt_val)
                    if pd.notna(gt_val) and str(gt_val).strip() != ""
                    else ""
                )
                model_str = str(model_val) if pd.notna(model_val) else ""

                gt_cell = ws_compare.cell(row, col, gt_str)
                if is_calculated:
                    gt_cell.fill = calculated_fill

                ws_compare.cell(row, col + 1, model_str)

                is_match, match_type = are_values_equivalent(gt_val, model_val, field)
                match_cell = ws_compare.cell(row, col + 2, "✓" if is_match else "✗")

                if is_match:
                    if match_type == "fuzzy":
                        # Fuzzy match (company name abbreviations, tax format)
                        match_cell.fill = fuzzy_match_fill  # Light orange
                    elif match_type == "zero_equivalent":
                        # Empty vs 0 equivalence
                        match_cell.fill = zero_match_fill  # Light blue
                    elif match_type == "numeric":
                        # Numeric match but different format
                        match_cell.fill = numeric_match_fill  # Light green
                    else:
                        # Exact match
                        match_cell.fill = match_fill  # Green
                else:
                    match_cell.fill = mismatch_fill  # Red

                col += 3
            row += 1

    # Adjust widths
    ws_compare.column_dimensions["A"].width = 40
    start_col = 3 if "invoice_type" in gt_df.columns else 2
    for col_idx in range(start_col, len(EVAL_FIELDS) * 3 + start_col):
        col_letter = ws_compare.cell(1, col_idx).column_letter
        ws_compare.column_dimensions[col_letter].width = 20

    return overall_accuracy, overall_precision, overall_recall, overall_f1
