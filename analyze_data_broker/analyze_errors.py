#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Analyze errors from broker accuracy Excel reports
Generate detailed error analysis report
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from collections import defaultdict

BASE_DIR = Path(__file__).parent
REPORT_DIR = BASE_DIR / "danh_gia_ket_qua" / "2026_02_04"
OUTPUT_FILE = BASE_DIR / "report_analysis_error.md"

REPORTS = [
    ("Contract Note", "accuracy_report_contract_note.xlsx"),
    ("Dividend Advice", "accuracy_report_dividend.xlsx"),
    ("FX Trade", "accuracy_report_fx_trade.xlsx"),
    ("Interest Payment", "accuracy_report_interest_payment.xlsx"),
    ("Trade Confirmation", "accuracy_report_trade_confirmation.xlsx"),
]


def analyze_excel_report(excel_file):
    """Analyze errors from Excel report"""
    wb = load_workbook(excel_file, data_only=True)
    ws_summary = wb["Accuracy Summary"]
    ws_compare = wb["Field Comparison"]

    # Read summary metrics
    field_metrics = []
    for row_idx in range(2, ws_summary.max_row + 1):
        field = ws_summary.cell(row_idx, 1).value
        if not field or "OVERALL" in str(field):
            break

        total_gt = ws_summary.cell(row_idx, 2).value or 0
        total_model = ws_summary.cell(row_idx, 3).value or 0
        correct = ws_summary.cell(row_idx, 4).value or 0
        accuracy = ws_summary.cell(row_idx, 5).value

        if isinstance(accuracy, str):
            accuracy = float(accuracy.replace("%", ""))

        error_count = total_gt - correct
        error_rate = (error_count / total_gt * 100) if total_gt > 0 else 0

        field_metrics.append(
            {
                "field": field,
                "total_gt": int(total_gt),
                "correct": int(correct),
                "errors": int(error_count),
                "error_rate": float(error_rate),
                "accuracy": float(accuracy),
            }
        )

    # Read overall metrics
    overall_metrics = None
    for row_idx in range(1, ws_summary.max_row + 1):
        cell_value = ws_summary.cell(row_idx, 1).value
        if cell_value and "OVERALL SUMMARY" in str(cell_value):
            overall_metrics = {
                "total_gt": ws_summary.cell(row_idx, 2).value or 0,
                "total_model": ws_summary.cell(row_idx, 3).value or 0,
                "correct": ws_summary.cell(row_idx, 4).value or 0,
                "accuracy": ws_summary.cell(row_idx, 5).value,
                "precision": ws_summary.cell(row_idx, 6).value,
                "recall": ws_summary.cell(row_idx, 7).value,
                "f1_score": ws_summary.cell(row_idx, 8).value,
            }
            break

    return {
        "field_metrics": field_metrics,
        "overall_metrics": overall_metrics,
    }


def generate_report():
    """Generate error analysis report"""
    print("Analyzing broker accuracy reports...")

    all_results = {}

    for doc_type, report_file in REPORTS:
        report_path = REPORT_DIR / report_file

        if not report_path.exists():
            print(f"  Warning: {report_file} not found")
            continue

        print(f"  Analyzing: {doc_type}")
        all_results[doc_type] = analyze_excel_report(report_path)

    # Generate markdown report
    lines = []
    lines.append("# Phân Tích Lỗi & Nguyên Nhân (Broker Data Accuracy Analysis)")
    lines.append("")
    lines.append("**Cập nhật:** 2026-02-04")
    lines.append("")

    # 1. Overall Summary
    lines.append("## 1. Tổng Quan Kết Quả")
    lines.append("")
    lines.append(
        "| Document Type | Accuracy | Precision | Recall | F1-Score | Files | Nhận Xét |"
    )
    lines.append(
        "| :------------ | :------- | :-------- | :----- | :------- | :---- | :------- |"
    )

    for doc_type in [
        "Contract Note",
        "Dividend Advice",
        "FX Trade",
        "Interest Payment",
        "Trade Confirmation",
    ]:
        if doc_type not in all_results:
            continue

        metrics = all_results[doc_type]["overall_metrics"]
        if not metrics:
            continue

        acc = metrics["accuracy"]
        prec = metrics["precision"]
        rec = metrics["recall"]
        f1 = metrics["f1_score"]
        files = metrics["total_gt"]

        # Convert to float if string
        if isinstance(acc, str):
            acc = float(acc.replace("%", ""))
        else:
            acc = float(acc)
        if isinstance(prec, str):
            prec = float(prec.replace("%", ""))
        else:
            prec = float(prec)
        if isinstance(rec, str):
            rec = float(rec.replace("%", ""))
        else:
            rec = float(rec)
        if isinstance(f1, str):
            f1 = float(f1.replace("%", ""))
        else:
            f1 = float(f1)

        # Determine status
        if acc >= 95:
            status = "🟢 Excellent"
        elif acc >= 90:
            status = "✅ Very Good"
        elif acc >= 80:
            status = "🟡 Good"
        else:
            status = "🔴 Needs Improvement"

        lines.append(
            f"| **{doc_type}** | **{acc:.2f}%** | **{prec:.2f}%** | **{rec:.2f}%** | **{f1:.2f}%** | {files} | {status} |"
        )

    lines.append("")
    lines.append("---")
    lines.append("")

    # 2. Field-level analysis for each document type
    lines.append("## 2. Phân Tích Chi Tiết Theo Field")
    lines.append("")

    for doc_type in [
        "Contract Note",
        "Dividend Advice",
        "FX Trade",
        "Interest Payment",
        "Trade Confirmation",
    ]:
        if doc_type not in all_results:
            continue

        field_metrics = all_results[doc_type]["field_metrics"]
        if not field_metrics:
            continue

        # Sort by error rate descending
        sorted_fields = sorted(
            field_metrics, key=lambda x: x["error_rate"], reverse=True
        )

        lines.append(f"### {doc_type.upper()}")
        lines.append("")
        lines.append("| Field | Error % | Total | Correct | Errors | Nhận Xét |")
        lines.append("| :---- | :------ | :---- | :------ | :----- | :------- |")

        for field_data in sorted_fields:
            field = field_data["field"]
            error_rate = field_data["error_rate"]
            total = field_data["total_gt"]
            correct = field_data["correct"]
            errors = field_data["errors"]

            # Determine status
            if error_rate >= 30:
                status = "🔴 High error rate"
            elif error_rate >= 15:
                status = "🟡 Moderate errors"
            elif error_rate >= 5:
                status = "🟢 Acceptable"
            else:
                status = "🟢 Excellent"

            lines.append(
                f"| **{field}** | {error_rate:.1f}% | {total} | {correct} | {errors} | {status} |"
            )

        lines.append("")

    lines.append("---")
    lines.append("")

    # 3. Key Findings
    lines.append("## 3. Phát Hiện Chính")
    lines.append("")

    # Find worst performing document types
    lines.append("### A. Document Types Cần Cải Thiện")
    lines.append("")

    doc_accuracies = []
    for doc_type in all_results:
        metrics = all_results[doc_type]["overall_metrics"]
        if metrics:
            acc = metrics["accuracy"]
            if isinstance(acc, str):
                acc = float(acc.replace("%", ""))
            doc_accuracies.append((doc_type, acc))

    doc_accuracies.sort(key=lambda x: x[1])

    for doc_type, acc in doc_accuracies:
        if acc < 80:
            lines.append(f"**{doc_type}: {acc:.2f}%**")
            lines.append("")

            # Show top error fields
            field_metrics = all_results[doc_type]["field_metrics"]
            top_errors = sorted(
                field_metrics, key=lambda x: x["error_rate"], reverse=True
            )[:5]

            lines.append("Top error fields:")
            for field_data in top_errors:
                lines.append(
                    f"- **{field_data['field']}**: {field_data['error_rate']:.1f}% error rate ({field_data['errors']} errors)"
                )
            lines.append("")

    lines.append("### B. Fields Có Lỗi Cao Nhất")
    lines.append("")

    # Collect all fields across document types
    all_field_errors = []
    for doc_type in all_results:
        field_metrics = all_results[doc_type]["field_metrics"]
        for field_data in field_metrics:
            if field_data["error_rate"] >= 20:
                all_field_errors.append(
                    {
                        "doc_type": doc_type,
                        "field": field_data["field"],
                        "error_rate": field_data["error_rate"],
                        "errors": field_data["errors"],
                    }
                )

    all_field_errors.sort(key=lambda x: x["error_rate"], reverse=True)

    for item in all_field_errors[:10]:
        lines.append(
            f"- **{item['doc_type']} - {item['field']}**: {item['error_rate']:.1f}% ({item['errors']} errors)"
        )

    lines.append("")
    lines.append("---")
    lines.append("")

    # 4. Recommendations
    lines.append("## 4. Khuyến Nghị")
    lines.append("")
    lines.append("### Ưu Tiên Cao")
    lines.append("")

    for doc_type, acc in doc_accuracies:
        if acc < 80:
            lines.append(f"1. **Cải thiện {doc_type}** (hiện tại {acc:.2f}%)")
            lines.append("   - Review Excel report để xác định pattern lỗi")
            lines.append("   - Kiểm tra GT data quality")
            lines.append("   - Cải thiện model extraction logic")
            lines.append("")

    lines.append("### Ưu Tiên Trung Bình")
    lines.append("")
    lines.append("1. **Thêm fuzzy matching** cho các fields có variation:")
    lines.append("   - Company names")
    lines.append("   - Security names")
    lines.append("   - Account numbers")
    lines.append("")
    lines.append("2. **Chuẩn hóa date formats** để tránh format mismatch")
    lines.append("")
    lines.append("3. **Improve numeric parsing** cho amount fields")
    lines.append("")

    lines.append("---")
    lines.append("")
    lines.append("## 5. Next Steps")
    lines.append("")
    lines.append("1. Review Excel reports chi tiết cho từng document type")
    lines.append("2. Identify specific error patterns")
    lines.append("3. Update GT data nếu cần")
    lines.append("4. Improve model prompts/logic")
    lines.append("5. Re-run evaluation và so sánh kết quả")
    lines.append("")

    # Write to file
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    print(f"\nReport generated: {OUTPUT_FILE}")
    print(f"Total document types analyzed: {len(all_results)}")


if __name__ == "__main__":
    generate_report()
