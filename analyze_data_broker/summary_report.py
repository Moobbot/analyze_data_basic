#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Summary report for all broker accuracy evaluations
Reads metrics from Excel reports
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

BASE_DIR = Path(__file__).parent
REPORT_DIR = BASE_DIR / "danh_gia_ket_qua" / "2026_02_04"

# Report files
REPORTS = [
    ("Contract Note", "accuracy_report_contract_note.xlsx"),
    ("Dividend Advice", "accuracy_report_dividend.xlsx"),
    ("FX Trade", "accuracy_report_fx_trade.xlsx"),
    ("Interest Payment", "accuracy_report_interest_payment.xlsx"),
    ("Trade Confirmation", "accuracy_report_trade_confirmation.xlsx"),
]


def read_overall_metrics(excel_file):
    """Read overall metrics from Excel report"""
    try:
        wb = load_workbook(excel_file, data_only=True)
        ws = wb["Accuracy Summary"]

        # Find OVERALL SUMMARY row
        for row_idx in range(1, ws.max_row + 1):
            cell_value = ws.cell(row_idx, 1).value
            if cell_value and "OVERALL SUMMARY" in str(cell_value):
                # Read metrics from this row
                total_gt = ws.cell(row_idx, 2).value or 0
                total_model = ws.cell(row_idx, 3).value or 0
                correct = ws.cell(row_idx, 4).value or 0
                accuracy = ws.cell(row_idx, 5).value
                precision = ws.cell(row_idx, 6).value
                recall = ws.cell(row_idx, 7).value
                f1_score = ws.cell(row_idx, 8).value

                # Convert to float if string
                if isinstance(accuracy, str):
                    accuracy = float(accuracy.replace("%", ""))
                if isinstance(precision, str):
                    precision = float(precision.replace("%", ""))
                if isinstance(recall, str):
                    recall = float(recall.replace("%", ""))
                if isinstance(f1_score, str):
                    f1_score = float(f1_score.replace("%", ""))

                return {
                    "files": int(total_gt),
                    "accuracy": float(accuracy),
                    "precision": float(precision),
                    "recall": float(recall),
                    "f1_score": float(f1_score),
                }

        return None
    except Exception as e:
        print(f"  Error reading {excel_file.name}: {e}")
        return None


def main():
    print("=" * 80)
    print("BROKER ACCURACY EVALUATION - FINAL SUMMARY")
    print("=" * 80)
    print()

    results = []
    total_files = 0

    for doc_type, report_file in REPORTS:
        report_path = REPORT_DIR / report_file

        if not report_path.exists():
            print(f"Warning: {report_file} not found")
            results.append((doc_type, 0, None, None, None, None))
            continue

        metrics = read_overall_metrics(report_path)

        if metrics:
            results.append(
                (
                    doc_type,
                    metrics["files"],
                    metrics["accuracy"],
                    metrics["precision"],
                    metrics["recall"],
                    metrics["f1_score"],
                )
            )
            total_files += metrics["files"]
        else:
            results.append((doc_type, 0, None, None, None, None))

    # Print table
    print(
        f"{'Document Type':<25} {'Files':<8} {'Accuracy':<10} {'Precision':<10} {'Recall':<10} {'F1-Score':<10}"
    )
    print("-" * 80)

    for doc_type, files, acc, prec, rec, f1 in results:
        if acc is None:
            print(
                f"{doc_type:<25} {files:<8} {'N/A':<10} {'N/A':<10} {'N/A':<10} {'N/A':<10}"
            )
        else:
            print(
                f"{doc_type:<25} {files:<8} {acc:<10.2f} {prec:<10.2f} {rec:<10.2f} {f1:<10.2f}"
            )

    print("=" * 80)
    print()

    # Calculate average (excluding N/A)
    valid_results = [
        (acc, prec, rec, f1) for _, _, acc, prec, rec, f1 in results if acc is not None
    ]

    if valid_results:
        avg_acc = sum(r[0] for r in valid_results) / len(valid_results)
        avg_prec = sum(r[1] for r in valid_results) / len(valid_results)
        avg_rec = sum(r[2] for r in valid_results) / len(valid_results)
        avg_f1 = sum(r[3] for r in valid_results) / len(valid_results)

        print("AVERAGE METRICS:")
        print(f"  Accuracy:  {avg_acc:.2f}%")
        print(f"  Precision: {avg_prec:.2f}%")
        print(f"  Recall:    {avg_rec:.2f}%")
        print(f"  F1-Score:  {avg_f1:.2f}%")
        print()

    print("KEY FINDINGS:")
    print()

    # Sort by accuracy
    sorted_results = sorted(
        [(dt, acc) for dt, _, acc, _, _, _ in results if acc is not None],
        key=lambda x: x[1],
        reverse=True,
    )

    if sorted_results:
        print("  BEST PERFORMANCE:")
        for doc_type, acc in sorted_results[:2]:
            status = (
                "Excellent!" if acc >= 95 else "Very Good!" if acc >= 90 else "Good!"
            )
            print(f"    - {doc_type}: {acc:.2f}% ({status})")
        print()

        if len(sorted_results) > 2:
            print("  NEEDS IMPROVEMENT:")
            for doc_type, acc in sorted_results[-2:]:
                if acc < 80:
                    print(
                        f"    - {doc_type}: {acc:.2f}% (Review Excel report for details)"
                    )
            print()

    print(f"  TOTAL FILES EVALUATED: {total_files}")
    print()
    print("=" * 80)
    print()
    print(f"All reports available in: {REPORT_DIR}")
    print()


if __name__ == "__main__":
    main()
