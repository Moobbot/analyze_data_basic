#!/usr/bin/env python3
"""
Common utilities for broker data accuracy evaluation
"""

import pandas as pd
import re
from openpyxl.styles import Font, PatternFill


# Field lists for different document types
CONTRACT_NOTE_FIELDS = [
    "Client name",
    "Name/ Security",
    "Securities ID",
    "Transaction type",
    "Trade date",
    "Settlement date",
    "Quantity",
    "Foreign Unit Price",
    "Foreign Gross Consideration",
    "Foreign Net Consideration",
    "Net Consideration",
    "Currency",
    "Account no.",
    "Accrued Interest",
    "Exec Commission",
    "Research Commission",
    "Total Commission",
    "Local Fee",
    "Local Tax",
    "Stamp Duty",
    "Foreign GST",
    "GST equivalent in SGD",
    "GST ON (SR)",
]

DIVIDEND_FIELDS = [
    "Client name",
    "Name/ Security",
    "Securities ID",
    "Ex-Date",
    "Payment Date",
    "Currency",
    "Units",
    "Dividend Rate",
    "WHT Rate",
    "Gross Dividend Amount (Local)",
    "WHT Amount",
    "Net Dividend Amount (Local)",
    "Net consideration",
    "Account no.",
]

FX_TRADE_FIELDS = [
    "Client name",
    "Transaction type",
    "Trade date",
    "Settlement date",
    "Rate",
    "Currency Buy",
    "Amount Buy",
    "Currency Sell",
    "Amount Sell",
    "Account no. Buy",
    "Account no. Sell",
]

INTEREST_PAYMENT_FIELDS = [
    "Account no.",
    "Currency",
    "Date",
    "Transaction type",
    "Reference",
    "Amounts",
    "Value date",
    "Balances",
]

TRADE_CONFIRMATION_FIELDS = [
    "Client name",
    "Name/ Security",
    "Securities ID",
    "Transaction type",
    "Trade date",
    "Settlement date",
    "Quantity",
    "Foreign Unit Price",
    "Foreign Gross Consideration",
    "Foreign Net Consideration",
    "Net Consideration",
    "Currency",
    "Account no.",
]


def normalize_filename(filename):
    """Normalize filename for matching"""
    if pd.isna(filename):
        return ""

    # Remove extension
    name = str(filename).strip()
    if name.lower().endswith(".pdf"):
        name = name[:-4]
    if name.lower().endswith(".json"):
        name = name[:-5]

    return name.strip()


def normalize_value(value):
    """Normalize value for comparison"""
    if pd.isna(value):
        return ""

    value_str = str(value).strip()

    # Handle common variations
    if value_str.lower() in ["nan", "none", "null", "n/a"]:
        return ""

    return value_str


def safe_float(value):
    """Safely convert value to float"""
    if pd.isna(value) or value == "":
        return None

    try:
        # Remove commas and convert
        return float(str(value).replace(",", ""))
    except (ValueError, TypeError):
        return None


def normalize_date(date_str):
    """Normalize date to MM/DD/YYYY format"""
    if pd.isna(date_str) or str(date_str).strip() == "":
        return ""

    date_str = str(date_str).strip()

    # Already in correct format
    if re.match(r"^\d{2}/\d{2}/\d{4}$", date_str):
        return date_str

    # Try other common formats
    # Add more patterns as needed

    return date_str


def normalize_transaction_type(trans_type):
    """Normalize transaction type"""
    if pd.isna(trans_type) or str(trans_type).strip() == "":
        return ""

    # Uppercase and trim
    normalized = str(trans_type).strip().upper()

    # Handle common variations
    variations = {
        "PREPAYMENT FOR FUND SUBSCRIPTION": "PREPAYMENT FOR FUND SUBSCRIPTION",
        "ADJUSTMENT MAX. NOTIONAL": "ADJUSTMENT MAX. NOTIONAL",
        # Add more as needed
    }

    return variations.get(normalized, normalized)


def normalize_company_name(name):
    """Normalize company name for fuzzy matching"""
    if pd.isna(name) or str(name).strip() == "":
        return ""

    name = str(name).strip().lower()

    # Common abbreviations mapping
    replacements = {
        r"\bltd\.?\b": "limited",
        r"\bpte\.?\b": "private",
        r"\bco\.?\b": "company",
        r"\bcorp\.?\b": "corporation",
        r"\binc\.?\b": "incorporated",
        r"\bllc\.?\b": "limited liability company",
        r"\bllp\.?\b": "limited liability partnership",
        r"\bs\.?a\.?\b": "sociedad anonima",
        r"\bspc\.?\b": "segregated portfolio company",
    }

    for pattern, replacement in replacements.items():
        name = re.sub(pattern, replacement, name)

    # Remove extra spaces
    name = re.sub(r"\s+", " ", name).strip()

    return name


def are_values_equivalent(val1, val2, field_name=None):
    """
    Check if two values are equivalent
    Returns: (is_match: bool, match_type: str)
    """
    # Handle empty values
    if pd.isna(val1) and pd.isna(val2):
        return True, "exact"

    str1 = str(val1).strip() if not pd.isna(val1) else ""
    str2 = str(val2).strip() if not pd.isna(val2) else ""

    if str1 == "" and str2 == "":
        return True, "exact"

    # Check if both are zero-equivalent
    def is_zero_value(s):
        if s == "":
            return True
        try:
            num = float(s.replace(",", "").replace("-", ""))
            return num == 0.0
        except (ValueError, TypeError):
            return False

    is_zero1 = is_zero_value(str1)
    is_zero2 = is_zero_value(str2)

    if is_zero1 and is_zero2:
        if (str1 == "" and str2 != "") or (str1 != "" and str2 == ""):
            return True, "zero_equivalent"
        if str1 != str2:
            return True, "zero_equivalent"
        return True, "exact"

    # Fuzzy matching for Client Name
    if field_name and "Client name" in field_name:
        norm1 = normalize_company_name(val1)
        norm2 = normalize_company_name(val2)
        if norm1 == norm2:
            if str1.lower() != str2.lower():
                return True, "fuzzy"
            return True, "exact"

    # Exact match
    if str1 == str2:
        return True, "exact"

    # Numeric match (for numbers with different formatting)
    num1 = safe_float(val1)
    num2 = safe_float(val2)

    if num1 is not None and num2 is not None:
        if abs(num1 - num2) < 0.01:  # Tolerance for floating point
            if str1 != str2:
                return True, "numeric"
            return True, "exact"

    # Case-insensitive match for text fields
    if str1.lower() == str2.lower():
        return True, "fuzzy"

    return False, "mismatch"


def calculate_field_accuracy(field, gt_df, model_df, matched_files):
    """Calculate accuracy metrics for a specific field"""

    # Filter to matched files
    gt_matched = gt_df[gt_df["_filename_normalized"].isin(matched_files)]
    model_matched = model_df[model_df["_filename_normalized"].isin(matched_files)]

    total_gt = len(gt_matched)
    total_model = len(model_matched)
    correct = 0

    for filename in matched_files:
        gt_row = gt_matched[gt_matched["_filename_normalized"] == filename]
        model_row = model_matched[model_matched["_filename_normalized"] == filename]

        if len(gt_row) == 0 or len(model_row) == 0:
            continue

        gt_val = gt_row.iloc[0].get(field)
        model_val = model_row.iloc[0].get(field)

        is_match, _ = are_values_equivalent(gt_val, model_val, field)
        if is_match:
            correct += 1

    # Calculate metrics
    accuracy = (correct / total_gt * 100) if total_gt > 0 else 0
    precision = (correct / total_model * 100) if total_model > 0 else 0
    recall = (correct / total_gt * 100) if total_gt > 0 else 0
    f1_score = (
        (2 * precision * recall / (precision + recall))
        if (precision + recall) > 0
        else 0
    )

    return {
        "field": field,
        "total_gt": total_gt,
        "total_model": total_model,
        "correct": correct,
        "accuracy": accuracy,
        "precision": precision,
        "recall": recall,
        "f1_score": f1_score,
    }


def create_excel_report(wb, gt_df, model_df, results, matched_files, field_list):
    """Create Excel report with accuracy summary and field comparison"""

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Styling
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True)
    match_fill = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )
    mismatch_fill = PatternFill(
        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
    )
    numeric_match_fill = PatternFill(
        start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
    )
    zero_match_fill = PatternFill(
        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
    )
    fuzzy_match_fill = PatternFill(
        start_color="FFE6CC", end_color="FFE6CC", fill_type="solid"
    )

    # Sheet 1: Accuracy Summary
    ws_summary = wb.create_sheet("Accuracy Summary")

    headers = [
        "Field",
        "Total GT",
        "Total Model",
        "Correct",
        "Accuracy (%)",
        "Precision (%)",
        "Recall (%)",
        "F1-Score (%)",
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws_summary.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, result in enumerate(results, start=2):
        ws_summary.cell(row_idx, 1, result["field"])
        ws_summary.cell(row_idx, 2, result["total_gt"])
        ws_summary.cell(row_idx, 3, result["total_model"])
        ws_summary.cell(row_idx, 4, result["correct"])
        ws_summary.cell(row_idx, 5, f"{result['accuracy']:.2f}")
        ws_summary.cell(row_idx, 6, f"{result['precision']:.2f}")
        ws_summary.cell(row_idx, 7, f"{result['recall']:.2f}")
        ws_summary.cell(row_idx, 8, f"{result['f1_score']:.2f}")

    # Overall Summary
    summary_row = len(results) + 3
    ws_summary.cell(summary_row, 1, "OVERALL SUMMARY").font = Font(bold=True)

    total_gt = sum(r["total_gt"] for r in results)
    total_model = sum(r["total_model"] for r in results)
    total_correct = sum(r["correct"] for r in results)

    overall_accuracy = (total_correct / total_gt * 100) if total_gt > 0 else 0
    overall_precision = (total_correct / total_model * 100) if total_model > 0 else 0
    overall_recall = (total_correct / total_gt * 100) if total_gt > 0 else 0
    overall_f1 = (
        (2 * overall_precision * overall_recall / (overall_precision + overall_recall))
        if (overall_precision + overall_recall) > 0
        else 0
    )

    ws_summary.cell(summary_row, 2, total_gt)
    ws_summary.cell(summary_row, 3, total_model)
    ws_summary.cell(summary_row, 4, total_correct)
    ws_summary.cell(summary_row, 5, f"{overall_accuracy:.2f}")
    ws_summary.cell(summary_row, 6, f"{overall_precision:.2f}")
    ws_summary.cell(summary_row, 7, f"{overall_recall:.2f}")
    ws_summary.cell(summary_row, 8, f"{overall_f1:.2f}")

    # Add color legend
    note_row = summary_row + 2
    ws_summary.cell(note_row, 1, "COLOR LEGEND (Field Comparison Sheet):").font = Font(
        bold=True, italic=True
    )

    legend_items = [
        ("🟢 GREEN", "Exact match", match_fill),
        ("🟡 LIGHT GREEN", "Numeric match (different format)", numeric_match_fill),
        ("🔵 LIGHT BLUE", "Zero-equivalent match", zero_match_fill),
        ("🟠 LIGHT ORANGE", "Fuzzy match (case-insensitive)", fuzzy_match_fill),
        ("🔴 RED", "Mismatch / Error", mismatch_fill),
    ]

    current_row = note_row + 1
    for label, description, fill_color in legend_items:
        cell_label = ws_summary.cell(current_row, 1, label)
        cell_label.font = Font(bold=True, size=9)

        cell_desc = ws_summary.cell(current_row, 2, description)
        cell_desc.font = Font(size=9)

        cell_sample = ws_summary.cell(current_row, 5, "Sample")
        cell_sample.fill = fill_color
        cell_sample.font = Font(size=9)

        current_row += 1

    ws_summary.column_dimensions["A"].width = 30
    for col in ["B", "C", "D", "E", "F", "G", "H"]:
        ws_summary.column_dimensions[col].width = 15

    # Sheet 2: Field Comparison
    ws_compare = wb.create_sheet("Field Comparison")

    col = 1
    ws_compare.cell(1, col, "Filename").fill = header_fill
    ws_compare.cell(1, col, "Filename").font = header_font
    col += 1

    for field in field_list:
        ws_compare.cell(1, col, f"GT_{field}").fill = header_fill
        ws_compare.cell(1, col, f"GT_{field}").font = header_font
        ws_compare.cell(1, col + 1, f"Model_{field}").fill = header_fill
        ws_compare.cell(1, col + 1, f"Model_{field}").font = header_font
        ws_compare.cell(1, col + 2, "Match").fill = header_fill
        ws_compare.cell(1, col + 2, "Match").font = header_font
        col += 3

    row = 2
    for filename in sorted(matched_files):
        gt_row_data = gt_df[gt_df["_filename_normalized"] == filename]
        model_row_data = model_df[model_df["_filename_normalized"] == filename]

        if len(gt_row_data) == 0 or len(model_row_data) == 0:
            continue

        gt_data = gt_row_data.iloc[0].to_dict()
        model_data = model_row_data.iloc[0].to_dict()

        col = 1
        ws_compare.cell(row, col, filename)
        col += 1

        for field in field_list:
            gt_val = gt_data.get(field)
            model_val = model_data.get(field)

            gt_str = str(gt_val) if pd.notna(gt_val) else ""
            model_str = str(model_val) if pd.notna(model_val) else ""

            ws_compare.cell(row, col, gt_str)
            ws_compare.cell(row, col + 1, model_str)

            is_match, match_type = are_values_equivalent(gt_val, model_val, field)
            match_cell = ws_compare.cell(row, col + 2, "✓" if is_match else "✗")

            if is_match:
                if match_type == "fuzzy":
                    match_cell.fill = fuzzy_match_fill
                elif match_type == "zero_equivalent":
                    match_cell.fill = zero_match_fill
                elif match_type == "numeric":
                    match_cell.fill = numeric_match_fill
                else:
                    match_cell.fill = match_fill
            else:
                match_cell.fill = mismatch_fill

            col += 3

        row += 1

    # Adjust widths
    ws_compare.column_dimensions["A"].width = 20
    for col_idx in range(2, len(field_list) * 3 + 2):
        col_letter = ws_compare.cell(1, col_idx).column_letter
        ws_compare.column_dimensions[col_letter].width = 20

    return overall_accuracy, overall_precision, overall_recall, overall_f1
